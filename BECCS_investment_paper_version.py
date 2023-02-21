## ROBUST DECISION MAKING (RDM) MODEL FOR BECCS
# Author: Oscar Stenström
# Date: 2023-02-16

# This program was developed to perform a Robust Decision Making (RDM) analysis of the investment 
# decision of deploying a bioenergy carbon capture and storage deployment (BECCS) facility. The 
# case studied is the planned BECCS facility of Stockholm Exergi (more info at www.beccs.se).
# The program utilizes Rhodium, an open-source Python library for RDM and exploratory modeling: 
# https://github.com/Project-Platypus/Rhodium
# To run the program, first follow the installation guide provided in the repository. Make sure 
# this file "BECCS_investment_paper_version.py" is in the root "Rhodium" folder.

## GLOSSARY
# NE       = negative emission, i.e. one tonne of atmospheric CO2 permanently removed and stored 
# CF       = cash flow
# p        = price
# dt       = "likely" change over time (high value => higher probability of the parameter increasing
#                                        low value => higher probability of the parameter decreasing)
# y        = year
# AUCTION  = a percentage of specific BECCS costs covered by a reversed auctions policy 
# yQUOTA   = a year when quota obligations of NEs could be imposed
# yEUint   = a year when NEs could be integrated in an EU scheme for carbon removal trading
# yBIOban  = a year when the usage of biomass as renewable energy could be severely restricted
# yCLAIM   = a year when the claiming and selling of NEs to the voluntary carbon market is allowed
# SOW      = state of the world, i.e. one specific combination of conditions (parameter values)
# -Invest- = when emphasized, Invest refers to the decision to invest in the BECCS facility
# -Wait-   = when emphasized, Wait refers to the decision to not invest in the BECCS facility

## MODEL BEGINS HERE
import math
import numpy as np
import numpy_financial as npf
from scipy.optimize import brentq as root
from rhodium import *
import pandas as pd
import csv
import openpyxl
import random 
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.patches as mpatches

## CREATE EMPTY EXCEL SHEET FOR RESULTS
RDM_results_excel = openpyxl.Workbook()
sheet = RDM_results_excel.active
sheet.title = "Results for each SOW"

## CONSTRUCT THE CALCULATION MODEL 
# The calculations below aim to quantify the annual prices of electricity, heat and NEs (i.e. the 
# revenues), as well as the different costs of the Invest and Wait strategies. This is done for 27 
# time slots, as the model runs from 2024 (year 0) to 2050 (year 26). These values are then used to
# calculate the Net Present Value (NPV) of both strategies, and Internal Rate of Return (IRR) (only 
# real solutions) of the Invest strategy. 

# One can set investment_decision = 1 if Invest is of interest, or investment_decision = 0 if Wait 
# is of interest. 

# One can set pelectricity_dt = [-1,0.4] depending on if trends of lower or higher electricity 
# prices are explored, or pheat_dt = [-0.5,0.7] depending on if trends of lower or higher heat 
# prices are explored. 
def BECCS_investment(investment_decision,
        # Set some nominal values as function inputs, as desired.
        pelectricity_2024 = 50,                                                    #[EUR/MWh]        
        pheat_2024 = 50,                                                           #[EUR/MWh]  
        pNE_2024 = 30,                                                             #[EUR/tCO2]  
        pETS_2024 = 80,                                                            #[EUR/tCO2]  
        pbiomass = 25,                                                             #[EUR/MWh]  
        pelectricity_dt = 0.4,          #-1 to 0.4,   sets likelihood of price increase/reduction
        pheat_dt = -0.5,                #-0.5 to 0.7, sets likelihood of price increase/reduction
        pNE_dt = 0.3,                   #-1 to +1,    sets likelihood of price increase/reduction
        pETS_dt = 0,                    #-1 to +1,    sets exponential increase in ETS prices

        Discount_rate = 0.06,                                                       #[-]   
        Learning_rate = 0.01,                                                       #[-]          
        CAPEX = 200*10**6,                                                          #[EUR]   
        OPEX_fixed = 20*10**6,                                                      #[EUR/year]   
        OPEX_variable = 44,                                                         #[EUR/tCO2]   
        Cost_transportation = 22,                                                   #[EUR/tCO2]   
        Cost_storage = 14.5,                                                        #[EUR/tCO2]      
        AUCTION = 0.5,                                                              #[-] 
        yQUOTA = 2035,
        yEUint = 2040,
        yBIOban = 2051,
        yCLAIM = 2026,
        ):

    # POWER PLANT OPERATING CONDITIONS:
    # Operating conditions are not modelled as uncertainties, as they are routinely managed by the power
    # plant operators. Refer to full article for data sources.
    Qbiomass_input = 362                                                            #[MW]        
    Wpower_output_wait = 110                                                        #[MW]
    Qheat_output_wait = 287                                                         #[MW]
    Wpower_output_invest = 53                                                       #[MW]
    Qheat_output_invest = 337                                                       #[MW]

    Operating_hours = 8760 * 0.7                                                    #[h], rule of thumb of ~70 % operating hours/year.
    CO2capture_rate = 0.3                                                           #[tCO2/MWh_biomass]
    CO2captured = CO2capture_rate*Qbiomass_input*Operating_hours*(1-0.05)           #[tCO2/year], about 5 % of CO2 is leaked across the value chain.
    OPEX_power_plant = 29000*Qbiomass_input + 0.5*Operating_hours*Qbiomass_input    #[EUR/year], see full article for these operational costs.

    # CALCULATE ENERGY/CO2 PRICES FOR THIS SOW:
    # Helping functions are used to construct pseudo-random energy and CO2 price projections. The 
    # probabilities of price increases/decreases are influenced by the _dt parameters. 
    pelectricity = find_penergy(pelectricity_2024, pelectricity_dt, pfloor=5,  proof=200, pmaximum_increase=40, pmaximum_decrease=-40)
    pheat = find_penergy(pheat_2024, pheat_dt, pfloor=48, proof=200, pmaximum_increase=10, pmaximum_decrease=-10) 
    pNE = find_pNE(pNE_2024, pNE_dt, pfloor=3, proof=500, pmaximum_increase=40, pmaximum_decrease=-40)
    pETS = find_pETS(pETS_2024, pETS_dt, pmaximum_increase=40, pmaximum_decrease=-40)
    pbiomass = pbiomass #Already defined, restated for clarity.
    # It is now possible to calculate NPV values!

    #CALCULATE CASH FLOWS BASED ON BOTH INVESTMENT DECISIONS:
    # (1) calculate NPV for not investing, i.e. Waiting:
    NPV_wait = 0
    CFvec_wait = []
    for t in range(0,27):
        CF = (Wpower_output_wait*pelectricity[t]+Qheat_output_wait*pheat[t]-Qbiomass_input*pbiomass)*Operating_hours - OPEX_power_plant   
        if (2023+t >= yBIOban):
            CF -= CO2captured*pETS[t]           #yBIOban represents a severe restriction of biomass usage, forcing the utility to pay for emission allowances for CO2 not captured.
        NPV_wait += CF/((1+Discount_rate)**t)
        CFvec_wait.append(CF)                   
    # Now NPV is known for the Wait strategy!

    # (2) calculate NPV (and IRR) for the Invest strategy:
    NPV_invest = 0
    CFvec = []
    CFvec_IRR = []                              #NOTE: For this analysis, NPV and IRR use different cashflows and are therefore not directly linked. 
    pNE_supported = []                          #In this vector we save pNE_max, which is the maximum CO2 price offered from any of the different support policy models.
    for t in range(0,2):                        #First two years we pay CAPEX, and we do not have revenues from NEs.
        CF = (Wpower_output_wait*pelectricity[t]+Qheat_output_wait*pheat[t]-Qbiomass_input*pbiomass)*Operating_hours - OPEX_power_plant - CAPEX/2 
        CFIRR = - CAPEX/2                       #IRR only includes incomes/costs generated by the CAPEX, and not the baseline revenues from heat and power.
        NPV_invest += CF/((1+Discount_rate)**t)

        CFvec.append(CF)
        CFvec_IRR.append(CFIRR)                 
        pNE_supported.append(pNE[t])            #The first two years, the maximum CO2 price is just the voluntary carbon market (VCM) price pNE(t).

    for t in range(2,27):
        CFenergy = (Wpower_output_invest*pelectricity[t]+Qheat_output_invest*pheat[t]-Qbiomass_input*pbiomass)*Operating_hours - OPEX_power_plant
        CFenergy_IRR = -((Wpower_output_wait-Wpower_output_invest)*pelectricity[t]+(Qheat_output_wait-Qheat_output_invest)*pheat[t])*Operating_hours #This is the energy penalty incurred by the CAPEX investment.

        Cost_specific = ( (OPEX_variable+Cost_transportation+Cost_storage)+OPEX_fixed/CO2captured )*(1-Learning_rate*(t-2))
        pNE_max = pNE[t]                        #Now, what is the maximum NE price we can sell to in this SOW? Answer: the highest of the VCM price, and prices achieved from uncertain policy support: 
        if 2024+t >= yQUOTA:                    #If quota obligations (yQUOTA) exist, NE price is assumed to be _at least_ equal to the specific cost.
            if Cost_specific > pNE_max:
                pNE_max = Cost_specific       
        if 2024+t >= yEUint:                    #If ETS integration (yEUint) exist, we sell to a price equivalent to ETS levels, if that price is higher.
            if pETS[t] > pNE_max:
                pNE_max = pETS[t]
        if 2024+t <= 2040:                      #Reversed auctions (AUCTION) reduces the _specific_ costs, until 2040.
            Cost_specific = Cost_specific * (1-AUCTION)
        
        # It is now possible to calculate cash flows from the maximum pNE offered, CFCO2.
        # However: we can not sell NEs (i.e. price is set to zero) if EU severely restricts biomass usage (yBIOban), or if we can't claim NEs (yCLAIM):
        if (2024+t < yBIOban) and (2024+t > yCLAIM): 
            CFCO2 = pNE_max*CO2captured - (Cost_specific*CO2captured)
        else:
            pNE_max = 0
            CFCO2 = pNE_max*CO2captured - (Cost_specific*CO2captured)
            
        # Now when cash flows from energy and CO2 is known, NPV of Investing can be calculated:
        CF = CFenergy+CFCO2
        NPV_invest += CF/((1+Discount_rate)**t)

        CFIRR = CFenergy_IRR+CFCO2              #Reminder: this cashflow for IRR is different than for NPV. 
        if (2023+t >= yBIOban):
            CFIRR += CO2captured*pETS[t]        #When calculating IRR of Investing, we sometimes avoid ETS costs.

        CFvec.append(CF)
        CFvec_IRR.append(CFIRR)
        pNE_supported.append(pNE_max)
    # Now NPV (and IRR) is known for the Invest strategy!

    ## DETERMINE REGRET BASED ON INVESTMENT DECISION:
    if investment_decision == 0:
        Regret = max(NPV_invest,NPV_wait) - NPV_wait

    if investment_decision == 1:                #NOTE: this is the decision in focus for the article.
        Regret = max(NPV_invest,NPV_wait) - NPV_invest

    ## CALCULATE OTHER INTERESTING PARAMETERS:
    pelectricity_mean = np.mean(pelectricity) 
    pheat_mean = np.mean(pheat)   
    pNE_mean = np.mean(pNE_supported)
    IRR = npf.irr(CFvec_IRR) 
    if math.isnan(IRR):
        IRR = 0

    return (pNE_mean, NPV_invest, Regret, NPV_wait, IRR, pelectricity_mean, pheat_mean)
#Now the calculation model is done!
    
## DEFINE RHODIUM MODEL
model = Model(BECCS_investment)
model.parameters = [Parameter("investment_decision"),

                    Parameter("pNE_2024"),  
                    Parameter("pNE_dt"),
                    Parameter("pelectricity_2024"),    
                    Parameter("pelectricity_dt"),
                    Parameter("pheat_2024"),  
                    Parameter("pheat_dt"),
                    Parameter("pbiomass"),
                    Parameter("pETS_2024"),
                    Parameter("pETS_dt"),

                    Parameter("Discount_rate"),
                    Parameter("CAPEX"),
                    Parameter("OPEX_fixed"),
                    Parameter("OPEX_variable"),
                    Parameter("Cost_transportation"),
                    Parameter("Cost_storage"),
                    Parameter("Learning_rate"),
                    
                    Parameter("AUCTION"),
                    Parameter("yQUOTA"),
                    Parameter("yEUint"),
                    Parameter("yBIOban"),
                    Parameter("yCLAIM")]

model.responses = [Response("pNE_mean", Response.INFO),
                   Response("NPV_invest", Response.MAXIMIZE),                 
                   Response("Regret", Response.MINIMIZE),
                   Response("NPV_wait", Response.MAXIMIZE),
                   Response("IRR", Response.MAXIMIZE),
                   Response("pelectricity_mean", Response.INFO),
                   Response("pheat_mean", Response.INFO),]

model.levers = [RealLever("investment_decision", 0, 1, length=2)] 

# For uncertainties, some are expanded ranges around values found in the literature, and some are assumed. Refer to full article.
model.uncertainties = [UniformUncertainty("pNE_2024", 20, 140),
                       UniformUncertainty("pNE_dt", -1, 1),
                       UniformUncertainty("pbiomass", 15, 35),
                       UniformUncertainty("pETS_2024", 60, 100),
                       UniformUncertainty("pETS_dt", -1,1),

                       UniformUncertainty("Discount_rate",         0.04, 0.10),
                       UniformUncertainty("CAPEX",      100*10**6, 300*10**6),
                       UniformUncertainty("OPEX_fixed",   18*10**6, 22*10**6),
                       UniformUncertainty("OPEX_variable",   33, 55),
                       UniformUncertainty("Cost_transportation",     17, 27),
                       UniformUncertainty("Cost_storage",     6, 23),
                       UniformUncertainty("Learning_rate",         0.0075, 0.0125),
                       
                       UniformUncertainty("AUCTION",         0, 1),
                       UniformUncertainty("yQUOTA",     2030, 2050),
                       UniformUncertainty("yEUint",       2035, 2050),
                       UniformUncertainty("yBIOban",       2030, 2050),
                       UniformUncertainty("yCLAIM",      2023, 2050),]

## DEFINE HELPING FUNCTIONS:
def find_pNE(pNE, pNE_dt, pfloor, proof, pmaximum_increase, pmaximum_decrease):
    #Translates pNE_dt to an alpha/beta value: 
    x = pNE_dt
    alpha = (1.4-0.6)/2*(x+1)+0.6   #These rows determine the "strength" of the _dt uncertainty. Can be experimented with to explore other price paths.
    beta  = (0.6-1.4)/2*(x+1)+1.4

    pNE_vec = []
    for t in range(0,27):
        pNE_vec.append(pNE)
        variance = random.betavariate(alpha*1.4, beta*1)    #This is how much the price change will be scaled, from this year to the next...

        if pNE + pmaximum_increase > proof:                 #... unless we hit a price roof/floor! Then adjust price accordingly.
            pvariance = pmaximum_decrease + ((proof-pNE)-pmaximum_decrease)*variance
        elif pNE + pmaximum_decrease < pfloor:
            pvariance = -(pNE-pfloor) + (pmaximum_increase-(-(pNE-pfloor)))*variance
        else:
            pvariance = pmaximum_decrease + (pmaximum_increase-pmaximum_decrease)*variance
        pNE += pvariance

    return pNE_vec

def find_penergy(pelectricity, pelectricity_dt, pfloor, proof, pmaximum_increase, pmaximum_decrease):
    #Works just like find_pNE()
    x = pelectricity_dt
    alpha = (1.7-0.3)/2*(x+1)+0.3 
    beta  = (0.3-1.7)/2*(x+1)+1.7

    pelectricity_vec = []
    for t in range(0,27):
        pelectricity_vec.append(pelectricity)
        variance = random.betavariate(alpha*1, beta*1) #*1.5 adds less random distribution, if desired.

        if pelectricity + pmaximum_increase > proof:
            pvariance = pmaximum_decrease + ((proof-pelectricity)-pmaximum_decrease)*variance
        elif pelectricity + pmaximum_decrease < pfloor:
            pvariance = -(pelectricity-pfloor) + (pmaximum_increase-(-(pelectricity-pfloor)))*variance
        else:
            pvariance = pmaximum_decrease + (pmaximum_increase-pmaximum_decrease)*variance
        pelectricity += pvariance

    return pelectricity_vec

def find_pETS(pETS_2024, pETS_dt, pmaximum_increase, pmaximum_decrease):
    #Works similar to find_pNE(), but pETS_dt is translated into a value between 0.6 and 1.5. 
    #This sets the strength of the exponential increase of pETS. Hard-coded values can be changed, if desired.
    x = (1.5-0.6)/2*(pETS_dt+1)+0.6 
    pETS_vec = [pETS_2024]
    for t in range(1,27):
        variance = random.betavariate(alpha=1, beta=1)
        pvariance = pmaximum_decrease + (pmaximum_increase-pmaximum_decrease)*variance

        pETS = pETS_2024*(1+0.06*x)**t + pvariance
        pETS_vec.append(pETS)

    return pETS_vec
    
## EVALUATE THE MODEL
random.seed(7)
policy = {"investment_decision" : 1}    #investment_decision = 1 means Invest and investment_decision = 0 means Wait.
SOWs = sample_lhs(model, 10000)         #10 000 SOWs are evaluated, and in the full article this is repeated for four energy price trends (i.e. by changing pelectricity_dt and pheat_dt). 
inputs = update(SOWs, policy)
model_results = evaluate(model, inputs)
model_results.save('RDM_raw_results.csv')

with open('RDM_raw_results.csv') as f:  # Now saving CSV results to the common Excel file.
    reader = csv.reader(f, delimiter=':')
    for row in reader:
        sheet.append(row)



## BELOW IS A MIX OF RESULTS ANALYSIS AND PLOTTING
print("-------------BEGIN RESPONSE PLOTTING NOW-------------")
fig = scatter2d(model, model_results, x="NPV_wait", y="NPV_invest", c="Regret")
fig.savefig("2_NPV_Regret.png")
fig = scatter2d(model, model_results, x="pNE_mean", y="NPV_invest", c="Regret")
fig.savefig("2_NPV_pNE.png")
fig = scatter2d(model, model_results, x="pheat_mean", y="pelectricity_mean", c="Regret")
fig.savefig("2_NPV_penergy.png")
fig = scatter2d(model, model_results.find("IRR != 0"), x="pNE_mean", y="IRR", c="Regret")
fig.savefig("2_IRR_pNE.png")
plt.clf()
pairs(model, model_results,brush=["Regret > 0", "Regret == 0"])
plt.savefig("2_Responses_pair.png")



print("-------------BEGIN ROBUSTNESS ANALYSIS NOW-------------")
# How robust is Invest and Wait, using the satisficing (absolute) domain criteria?
invest_satisficing = model_results.find("NPV_invest>0")
print("Investing is satisficing in ", len(invest_satisficing)," SOWs")
wait_satisficing   = model_results.find("NPV_wait>0")
print("Waiting is satisficing in ", len(wait_satisficing)," SOWs")

# How robust is Invest and Wait, using the satisficing (relative and absolute) domain criteria?
invest_satisficing = invest_satisficing.find("NPV_invest>NPV_wait")
print("Investing is _relative_ satisficing in ", len(invest_satisficing)," SOWs")
wait_satisficing   = wait_satisficing.find("NPV_invest<NPV_wait") 
print("Waiting is _relative_ satisficing in ", len(wait_satisficing)," SOWs")

# How robust is Invest and Wait, using the Savage criteria, i.e. to Min(Max(Regret))?
invest_regret_vec = []
wait_regret_vec = []
for SOW in model_results:
    invest_regret = max(SOW["NPV_invest"] , SOW["NPV_wait"]) - SOW["NPV_invest"]    #Finding Regret again from model results.
    wait_regret   = max(SOW["NPV_invest"] , SOW["NPV_wait"]) - SOW["NPV_wait"]      #Finding Regret again from model results.
    invest_regret_vec.append(invest_regret)
    wait_regret_vec.append(wait_regret)
print("Investing has maximum regret ", max(invest_regret_vec)," EUR")
print("Waiting has maximum regret ",   max(wait_regret_vec)," EUR")




print("-------------BEGIN SCENARIO DISCOVERY NOW-------------")
# The scenario discovery produces ranges of uncertainties (i.e. scenarios) where Invest performs well (i.e. have Regret = 0).
classification = model_results.apply("'Reliable' if Regret == 0 else 'Unreliable'")
cart_results = Cart(model_results, classification, include=model.uncertainties.keys(), min_samples_leaf=50)
# c.show_tree()
cart_results.save("4_CART_tree.png")
node_list = cart_results.print_tree(coi="Reliable") # NOTE: in the classification.py file of rhodium, the print_tree() function was edited.
                                                    # These edits basically just save the scenario nodes of the CART tree into a list, and return this list.

# Save discovered scenarios (in the node_list) to a CART excel sheet:
sheet = RDM_results_excel.create_sheet("CART_results")
RDM_results_excel.active = RDM_results_excel["CART_results"] 
sheet['A1'] = "Node nr"
sheet['B1'] = "Class"
sheet['C1'] = "Density"
sheet['D1'] = "Coverage"
sheet['E1'] = "Rule(s)"
for i, node in enumerate(node_list):
    sheet.cell( row=i+2, column=1 ).value = node["Node"]
    sheet.cell( row=i+2, column=2 ).value = node["Class"]
    sheet.cell( row=i+2, column=3 ).value = node["Density"]
    sheet.cell( row=i+2, column=4 ).value = node["Coverage"]
    if "Rules" in node.keys():
        for j, rule in enumerate(node["Rules"]): # "Rules" are ranges of uncertainties.
            sheet.cell( row=i+2, column=j+5 ).value = rule        

# The Rules (uncertainty ranges) of a scenario node of interest (as found in the CART_results sheet) can be illustrated.
# This is done by drawing a scenario rectangle representing these uncertainty ranges. The resulting "box" then graphically 
# represents a discovered scenario. The drawing is hard coded and can be changed as desired, depending on the scenario of interest.
fig = scatter2d(model, model_results, x="yCLAIM", y="pNE_dt", c="Regret")
scenario_area=mpatches.Rectangle((2023, 0.047607),(2032.198914-2023),1-(0.047607), 
                        fill=False,
                        color="crimson",
                       linewidth=3)
                       #facecolor="red")
plt.gca().add_patch(scenario_area)
fig.savefig("4_CART_area.png")



print("-------------BEGIN SENSITIVITY ANALYSIS NOW-------------")
# The sensitivity analysis indicates what uncertainties drive Regret. Using Sobols method, this is measured in 1st, 2nd and total order sensitivity indices.
# The article uses nsamples = 500 000, but for fast model evaluations nsamples = 10 000 can be used.
sobol_result = sa(model, "Regret", policy=policy, method="sobol", nsamples=10000) 
print(sobol_result) 

# Save Sobol results to a CART excel sheet:
sheet = RDM_results_excel.create_sheet("Sobol_results")
RDM_results_excel.active = RDM_results_excel["Sobol_results"] 
sheet['A1'] = "Uncertainty"
sheet['B1'] = "S1"
sheet['C1'] = "S1 (confidence interval)"
sheet['D1'] = "ST"
sheet['E1'] = "ST (confidence interval)"
i = 2
for name in model.uncertainties.keys():
    sheet.cell( row=i, column=1 ).value = name
    sheet.cell( row=i, column=2 ).value = sobol_result["S1"][name]
    sheet.cell( row=i, column=3 ).value = sobol_result["S1_conf"][name]
    sheet.cell( row=i, column=4 ).value = sobol_result["ST"][name]
    sheet.cell( row=i, column=5 ).value = sobol_result["ST_conf"][name]
    # sheet.cell( row=i, column=6 ).value = sobol_result["S2"][name] #Interaction effects are difficult to save.
    # sheet.cell( row=i, column=7 ).value = sobol_result["S2_conf"][name]
    i+=1
RDM_results_excel.save('RDM_processed_results.xlsx')

plt.clf()
fig = sobol_result.plot_sobol(radSc=1.9, widthSc=0.7, threshold=0.015, 
                        groups={"Commodity prices" : ["pNE_2024","pNE_dt", "pbiomass","pETS_2024","pETS_dt"],
                                "BECCS valuations" : ["Discount_rate", "CAPEX","OPEX_fixed","OPEX_variable","Cost_transportation","Cost_storage","Learning_rate"],
                                "Policy states" : ["AUCTION","yEUint", "yQUOTA","yBIOban","yCLAIM"]})
fig.savefig("3_Sobol_spider.png")

# Below one can plot the critical uncertainties (i.e. with high total sensitivity indices), to see how these affect Regret.
fig = scatter2d(model, model_results, x="yCLAIM", y="yBIOban", c="Regret")
fig.savefig("3_Sobol_Us.png")

print("\n-------------END OF MODEL-------------")
print("\nNOTE: you can comment out the CART node prints and the Sobol sensitivity analysis prints, as these anyway are saved to the RDM_processed_results file. Doing this makes it easier to see the robustness results, which are printed in the terminal.")