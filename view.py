"""ROBUST DECISION MAKING (RDM) MODEL FOR BECCS

This program was developed to perform a Robust Decision Making (RDM) analysis of the investment
decision of deploying a bioenergy carbon capture and storage deployment (BECCS) facility. The
case studied is the planned BECCS facility of Stockholm Exergi (more info at www.beccs.se).
The program utilizes Rhodium, an open-source Python library for RDM and exploratory modeling:
https://github.com/Project-Platypus/Rhodium
To run the program, first follow the installation guide provided in the repository. Make sure
this file "BECCS_investment_paper_version.py" is in the root "Rhodium" folder.
"""
__author__ = "Oscar Stenström"
__date__ = "2023-06-05"

from scipy.optimize import brentq as root
from rhodium import scatter2d, Cart, pairs, DataSet, Model, joint
import csv
import openpyxl
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from typing import List, Dict
from PIL import Image


def plot_results(model: Model, model_results: DataSet):
    ## BELOW IS A MIX OF RESULTS ANALYSIS AND PLOTTING
    print("-------------BEGIN RESPONSE PLOTTING NOW-------------")
    fig = scatter2d(model, model_results, x="NPV_wait", y="NPV_invest", c="Regret")
    fig.savefig("2_NPV_Regret.png", dpi=600)
    
    joint(model, model_results, x="NPV_wait", y="NPV_invest", color="turquoise")
    plt.savefig("2_NPV_distr.png", dpi=600)
    plt.clf()

    fig = scatter2d(model, model_results, x="pNE_supported", y="NPV_invest", c="Regret")
    fig.savefig("2_NPV_pNE.png", dpi=600)
    
    fig = scatter2d(model, model_results, x="Cost_specific", y="pNE_supported", c="Regret")
    fig.savefig("2_pNE_Costs.png", dpi=600)

    fig = scatter2d(model, model_results, x="Cost_specific", y="NPV_invest", c="Regret")
    fig.savefig("2_NPV_Costs.png", dpi=600)

    joint(model, model_results, x="Cost_specific", y="pNE_supported", color="turquoise")
    plt.savefig("2_Costs_distr.png", dpi=600)
    plt.clf()

    joint(model, model_results, x="pNE_mean", y="pNE_supported", color="turquoise")
    plt.savefig("2_pNE_distr.png", dpi=600)
    plt.clf()
    
    pairs(model, model_results, brush=["Regret > 0", "Regret == 0"])
    plt.savefig("2_Responses_Pair.png", dpi=600)
    plt.clf()
    
    n_successful   = len(model_results.find("Regret==0 and pNE_supported>120"))
    n_unsuccessful = len(model_results.find("Regret >0 and pNE_supported>120"))
    print( n_successful/(n_successful+n_unsuccessful)*100 , "% of scenarios have Regret = 0 when the NE price is above 120 EUR/t") #94%

def robustness_analysis(model_results: DataSet):
    """Prints robustness analytics to the terminal"""
    print("-------------BEGIN ROBUSTNESS ANALYSIS NOW-------------")
    # How robust is Invest and Wait, using the satisficing (absolute) domain criteria?
    invest_satisficing = model_results.find("NPV_invest>0")
    print("Investing is satisficing in ", len(invest_satisficing), " SOWs")
    wait_satisficing = model_results.find("NPV_wait>0")
    print("Waiting is satisficing in ", len(wait_satisficing), " SOWs")

    # How robust is Invest and Wait, using the satisficing (relative and absolute) domain criteria?
    invest_satisficing_relative = invest_satisficing.find("NPV_invest>NPV_wait")
    print("Investing is _relative_ satisficing in ", len(invest_satisficing_relative), " SOWs")
    wait_satisficing_relative = wait_satisficing.find("NPV_invest<NPV_wait")
    print("Waiting is _relative_ satisficing in ", len(wait_satisficing_relative), " SOWs")

    # How robust is Invest and Wait, using the Savage criteria, i.e. to Min(Max(Regret))?
    invest_regret_vec = []
    wait_regret_vec = []
    for SOW in model_results:
        invest_regret = (
            max(SOW["NPV_invest"], SOW["NPV_wait"]) - SOW["NPV_invest"]
        )  # Finding Regret again from model results.
        wait_regret = (
            max(SOW["NPV_invest"], SOW["NPV_wait"]) - SOW["NPV_wait"]
        )  # Finding Regret again from model results.
        invest_regret_vec.append(invest_regret)
        wait_regret_vec.append(wait_regret)
    print("Investing has maximum regret ", max(invest_regret_vec), " EUR")
    print("Waiting has maximum regret ", max(wait_regret_vec), " EUR")

    robustness_results = [len(invest_satisficing), len(wait_satisficing), len(invest_satisficing_relative), len(wait_satisficing_relative), max(invest_regret_vec), max(wait_regret_vec)]
    return robustness_results


def scenario_discovery(model: Model, model_results: DataSet) -> list:
    #NOW I WANT TO FIND THE INTERESTING SCENARIOS. WITH EXACT VALUES!
    #1) Find Regret=0. See what pop outs.


    # The scenario discovery produces ranges of uncertainties (i.e. scenarios) where Invest performs well (i.e. have Regret = 0).
    print("-------------BEGIN SCENARIO DISCOVERY NOW-------------")
    classification = model_results.apply("'Reliable' if (Regret == 0 and NPV_invest >= 0) else 'Unreliable'") 
    # Alternative classifications that can be applied depending on what analysis is made:
    # Regret == 0 and NPV_invest >= 0 # Drivers are then: yCLAM and pelectricity
    # pNE_supported-Cost_specific > 0
    # Regret != 0
    # Maybe find scenarios based on Exergi adaptation: pelectricity-pheat < 40 MWh. In such scenarios, what are the key drivers?

    cart_results = Cart(
        model_results,
        classification,
        include=model.uncertainties.keys(),
        min_samples_leaf=50,
    )
    cart_results.show_tree()
    cart_results.save("4_CART_tree.png")
    node_list = cart_results.print_tree(
        coi="Reliable"
    )  # NOTE: in the classification.py file of rhodium, the print_tree() function was edited.
    # These edits basically just save the scenario nodes of the CART tree into a list, and return this list.
    return node_list


def save_model_results(RDM_results_excel: openpyxl.Workbook, model_results: DataSet):
    model_results.save("RDM_raw_results.csv")
    sheet = RDM_results_excel.active  # typing: openpyxl.worksheet.worksheet.Worksheet
    sheet.title = "Results for each SOW"
    with open(
        "RDM_raw_results.csv"
    ) as f:  # Now saving CSV results to the common Excel file.
        reader = csv.reader(f, delimiter=":")
        for row in reader:
            sheet.append(row)

def save_robustness_analysis(robustness_results: list, RDM_results_excel: openpyxl.Workbook):
    sheet = RDM_results_excel.create_sheet("Robustness_results")
    RDM_results_excel.active = RDM_results_excel["Robustness_results"]
    sheet["A1"] = "Strategy"
    sheet["A2"] = "Invest"
    sheet["A3"] = "Wait"
    sheet["B1"] = "Satisficing [n SOWs]"
    sheet["B2"] = robustness_results[0]
    sheet["B3"] = robustness_results[1]
    sheet["C1"] = "Relative satisficing [n SOWs]"
    sheet["C2"] = robustness_results[2]
    sheet["C3"] = robustness_results[3]
    sheet["D1"] = "Maximum Regret [EUR]"
    sheet["D2"] = robustness_results[4]
    sheet["D3"] = robustness_results[5]
    RDM_results_excel.save("RDM_processed_results.xlsx")
   

def save_scenario_discovery(node_list: list, RDM_results_excel: openpyxl.Workbook):
    # Save discovered scenarios (in the node_list) to a CART excel sheet:
    sheet = RDM_results_excel.create_sheet("CART_results")
    RDM_results_excel.active = RDM_results_excel["CART_results"]
    sheet["A1"] = "Scenario node nr"
    sheet["B1"] = "Class"
    sheet["C1"] = "Density"
    sheet["D1"] = "Coverage"
    sheet["E1"] = "Rule(s)"
    for i, node in enumerate(node_list):
        sheet.cell(row=i + 2, column=1).value = node["Node"]
        sheet.cell(row=i + 2, column=2).value = node["Class"]
        sheet.cell(row=i + 2, column=3).value = node["Density"]
        sheet.cell(row=i + 2, column=4).value = node["Coverage"]
        if "Rules" in node.keys():
            for j, rule in enumerate(
                node["Rules"]
            ):  # "Rules" are ranges of uncertainties.
                sheet.cell(row=i + 2, column=j + 5).value = rule
    RDM_results_excel.save("RDM_processed_results.xlsx")

def plot_scenario_of_interest(model: Model, model_results: DataSet):
    # The Rules (uncertainty ranges) of a scenario node of interest (as found in the CART_results sheet) can be illustrated.
    # This is done by drawing a scenario rectangle representing these uncertainty ranges. The resulting "box" then graphically
    # represents a discovered scenario. The drawing is hard coded and can be changed as desired, depending on the scenario of interest.

    #-----------------The 1st scenario is plotted below----------
    # 2	Reliable	91.54516608426718	39.385508648844805	pelectricity_mean <= 109.776455	yCLAIM <= 2034.040527 (REGRET = 0)
    fig = scatter2d(model, model_results, x="yCLAIM", y="pelectricity_mean", c="Regret") 
    scenario_area = mpatches.Rectangle(
        (2024, 20),
        (2034 - 2024),
        110 - (20),
        fill=False,
        color="gold",
        linewidth=3,
    )
    # facecolor="red")
    plt.gca().add_patch(scenario_area)
    fig.savefig("4_Scenario_1.png", dpi=600) 
    plt.clf()

    #-----------------The 2nd scenario is plotted below----------
    # 1170	Reliable	75.04733607154301	61.26316790716477	pelectricity_mean > 81.914059	yCLAIM > 2034.032410 (REGRET != 0)
    fig = scatter2d(model, model_results, x="yCLAIM", y="pelectricity_mean", c="Regret")
    scenario_area = mpatches.Rectangle(
        (2034, 82),
        (2050 - 2034),
        160 - (82),
        fill=False,
        color="crimson",
        linewidth=3,
    )
    # facecolor="red")
    plt.gca().add_patch(scenario_area)
    fig.savefig("4_Scenario_2.png", dpi=600)
    plt.clf()

    #-----------------The 3rd scenario is plotted below----------
    # 206	Reliable	96.2998361551065	29.347147671951067	pNE_mean > 150.728317	yCLAIM <= 2030.209717 (REGRET = 0 AND Pelec>82)
    fig = scatter2d(model, model_results.find("pelectricity_mean>82"), x="yCLAIM", y="pNE_mean", c="Regret") 
    scenario_area = mpatches.Rectangle(
        (2024, 151),
        (2030 - 2024),
        300 - (151),
        fill=False,
        color="crimson",
        linewidth=3,
        linestyle = 'dashed',
    )
    # facecolor="red")
    plt.gca().add_patch(scenario_area)
    fig.savefig("4_Scenario_3.png", dpi=600)
    plt.clf()

    #-----------------The 4th scenario is plotted below----------
    # 95	Reliable	70.06802721088435	39.56005586592179	pETS_2050 > 232.763702	yBIOban <= 2034.975525 (REGRET = 0, AND Pelec>82 AND yCLAIM>2034)
    # 169	Reliable	88.32214765100672	22.97486033519553	pETS_2050 > 232.763702	AUCTION > 0.532030	yBIOban <= 2034.975525
    fig = scatter2d(model, model_results.find("pelectricity_mean>82 and yCLAIM>2034"), x="pETS_2050", y="yBIOban", c="Regret")
    scenario_area = mpatches.Rectangle(
        (233, 2030),
        (375 - 233),
        2035 - (2030),
        fill=False,
        color="crimson",
        linewidth=3,
        linestyle = 'dashed',
    )
    # facecolor="red")
    plt.gca().add_patch(scenario_area)
    fig.savefig("4_Scenario_4.png", dpi=600)
    plt.clf()

    #-----------------These rows can be used to combine plots into subplots----------
    # crop the right side of fig2 and fig3 by 5%
    img1 = Image.open("4_Scenario_1.png")
    img1_width, img1_height = img1.size
    img1_cropped = img1.crop((0, img1_width * 0.05, img1_width * 0.95, img1_height))
    img1_cropped.save("4_Scenario_1_cropped.png")

    img2 = Image.open("4_Scenario_2.png")
    img2_width, img2_height = img2.size
    img2_cropped = img2.crop((0, img2_width * 0.05, img2_width * 0.95, img2_height))
    img2_cropped.save("4_Scenario_2_cropped.png")

    img3 = Image.open("4_Scenario_3.png")
    img3_width, img3_height = img3.size
    img3_cropped = img3.crop((0, img3_width * 0.05, img3_width * 0.95, img3_height))
    img3_cropped.save("4_Scenario_3_cropped.png")

    img4 = Image.open("4_Scenario_4.png")
    img4_width, img4_height = img4.size
    img4_cropped = img4.crop((0, img4_width * 0.05, img4_width * 0.95, img4_height))
    img4_cropped.save("4_Scenario_4_cropped.png")

    # create a new figure and combine fig2 and fig3
    fig, axs = plt.subplots(2, 2, figsize=(12,9))
    axs[0,0].imshow(img1_cropped)
    axs[0,0].set_xlabel("(a)", fontsize=12)
    axs[0,0].axis("off")
    axs[0,1].imshow(img2_cropped)
    axs[0,1].set(xlabel="(b)")
    axs[0,1].axis("off")
    axs[1,0].imshow(img3_cropped)
    axs[1,0].set(xlabel="(c)")
    axs[1,0].axis("off")
    axs[1,1].imshow(img4_cropped)
    axs[1,1].set(xlabel="(d)")
    axs[1,1].axis("off")

    # adjust the layout by reducing the height spacing between subplots
    # plt.subplots_adjust(hspace=0)

    # adjust the layout and save the combined figure
    plt.tight_layout()
    plt.savefig("4_Scenarios_ALL.png", dpi=600)
    plt.clf()

def save_sensitivity_analysis(
    model: Model, sobol_result, RDM_results_excel: openpyxl.Workbook
):
    # Save Sobol results to a CART excel sheet:
    sheet = RDM_results_excel.create_sheet("Sobol_results")
    RDM_results_excel.active = RDM_results_excel["Sobol_results"]
    sheet["A1"] = "Uncertainty"
    sheet["B1"] = "S1"
    sheet["C1"] = "S1 (confidence interval)"
    sheet["D1"] = "ST"
    sheet["E1"] = "ST (confidence interval)"
    i = 2
    for name in model.uncertainties.keys():
        sheet.cell(row=i, column=1).value = name
        sheet.cell(row=i, column=2).value = sobol_result["S1"][name]
        sheet.cell(row=i, column=3).value = sobol_result["S1_conf"][name]
        sheet.cell(row=i, column=4).value = sobol_result["ST"][name]
        sheet.cell(row=i, column=5).value = sobol_result["ST_conf"][name]
        # sheet.cell( row=i, column=6 ).value = sobol_result["S2"][name] #Interaction effects are difficult to save.
        # sheet.cell( row=i, column=7 ).value = sobol_result["S2_conf"][name]
        i += 1
    RDM_results_excel.save("RDM_processed_results.xlsx")


def plot_sensitivity_analysis_results(sobol_result):
    #NOTE: can comment out the print, if desired.
    print(sobol_result) 
    plt.clf()
    fig = sobol_result.plot_sobol(
        radSc=1.9,
        widthSc=0.7,
        threshold=0.004, #0.015
        groups={
            "Commodity prices": [
                "pNE_mean",
                # "pNE_dt",
                # "pbiomass",
                "pETS_2050",
                # "pETS_dt",
                "pelectricity_mean",
                "pheat_mean",
            ],
            "BECCS financials": [
                "Discount_rate",
                # "CAPEX",
                # "OPEX_fixed",
                "OPEX_variable",
                # "Cost_transportation",
                # "Cost_storage",
                # "Learning_rate",
                # "Availability_factor,"
            ],
            "POLICY states": [
                "AUCTION", 
                # "yEUint", 
                # "yQUOTA", 
                "yBIOban", 
                "yCLAIM", 
                # "ySHOCK",
            ],
        },
    )
    fig.savefig("3_Sobol_spider1.png", dpi=600)
    plt.clf()
    fig = sobol_result.plot_sobol(
        radSc=1.9,
        widthSc=0.7,
        threshold=0.004, #0.015
        groups={
            " ": [
                "pNE_mean",
                # "pNE_dt",
                # "pbiomass",
                "pETS_2050",
                # "pETS_dt",
                "pelectricity_mean",
                "pheat_mean",
                "Discount_rate",
                # "CAPEX",
                # "OPEX_fixed",
                "OPEX_variable",
                # "Cost_transportation",
                # "Cost_storage",
                # "Learning_rate",
                # "Availability_factor,"
                "AUCTION", 
                # "yEUint", 
                # "yQUOTA", 
                "yBIOban", 
                "yCLAIM", 
                # "ySHOCK",
            ],
        },
    )
    fig.savefig("3_Sobol_spider2.png", dpi=600)


def plot_critical_uncertainties(model: Model, model_results: DataSet):
    # Below one can plot the critical uncertainties (i.e. with high total sensitivity indices), to see how these affect Regret.
    fig = scatter2d(model, model_results, x="yCLAIM", y="pelectricity_mean", c="Regret")
    fig.savefig("3_Sobol_Us1.png", dpi=600)

    fig = scatter2d(model, model_results, x="AUCTION", y="yBIOban", c="Regret")
    fig.savefig("3_Sobol_Us2.png", dpi=600)
